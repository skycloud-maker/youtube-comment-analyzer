"""Excel workbook writer for business users."""

from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# Columns to include in 댓글원문 / 대표댓글 exports (internal/technical columns excluded)
_COMMENT_EXPORT_COLUMNS: list[str] = [
    "video_id", "comment_id", "is_reply", "source_type",
    "author_display_name", "text_original", "like_count",
    "published_at", "updated_at", "reply_count", "total_reply_count",
    "text_length", "language_detected",
    "comment_quality", "comment_validity", "exclusion_reason", "removed_reason",
    "pii_flag", "pii_types",
    "product", "keyword", "region", "title", "channel_title", "video_url",
    "product_related", "product_target",
    "sentiment_label", "sentiment_score", "signal_strength",
    "classification_type", "confidence_level", "brand_label",
    "classification_reason", "insight_type", "parent_comment_text",
    "nlp_label", "nlp_confidence", "nlp_sentiment_reason",
    "nlp_topics", "nlp_topic_sentiments",
    "nlp_is_inquiry", "nlp_is_rhetorical",
    "nlp_summary", "nlp_user_wants", "nlp_keywords",
    "nlp_product_mentions", "nlp_core_points", "nlp_context_tags",
    "nlp_insight_summary", "nlp_sentiment_intensity",
    "journey_stage", "judgment_axes", "topic_label",
]

# (시트명, 컬럼명, 한국어 이름, 설명, 가능한 값)
_COLUMN_DESCRIPTIONS: list[tuple[str, str, str, str, str]] = [
    # ── 댓글원문 ──────────────────────────────────────────────────────────────
    ("댓글원문", "video_id", "영상 ID", "YouTube 영상 고유 식별자", "YouTube video ID 문자열"),
    ("댓글원문", "comment_id", "댓글 ID", "YouTube 댓글 고유 식별자", "YouTube comment ID 문자열"),
    ("댓글원문", "is_reply", "대댓글 여부", "최상위 댓글이면 False, 대댓글이면 True", "True / False"),
    ("댓글원문", "source_type", "출처 유형", "댓글 수집 경로 구분", "comment / reply"),
    ("댓글원문", "author_display_name", "작성자 이름", "댓글 작성자의 YouTube 표시 이름", "텍스트"),
    ("댓글원문", "text_original", "원문 댓글", "YouTube API에서 수집한 댓글 원본 텍스트", "텍스트"),
    ("댓글원문", "like_count", "좋아요 수", "댓글에 달린 좋아요 수 (engagement 지표)", "0 이상 정수"),
    ("댓글원문", "published_at", "작성 일시", "댓글이 게시된 날짜/시간 (UTC)", "ISO 8601 datetime"),
    ("댓글원문", "reply_count", "대댓글 수", "이 댓글에 달린 대댓글 수", "0 이상 정수"),
    ("댓글원문", "text_length", "댓글 길이", "원문 텍스트의 문자 수", "0 이상 정수"),
    ("댓글원문", "language_detected", "감지 언어", "텍스트에서 감지된 언어 코드", "ko / en / ja / zh 등"),
    ("댓글원문", "comment_quality", "댓글 품질", "텍스트 품질 자동 평가 결과", "high / medium / low"),
    ("댓글원문", "comment_validity", "유효성", "분석에 포함 가능한지 여부", "valid / excluded"),
    ("댓글원문", "exclusion_reason", "제외 사유", "유효성이 excluded일 때 제외된 이유", "spam / too_short / off_topic 등"),
    ("댓글원문", "pii_flag", "개인정보 포함 여부", "전화번호·이메일 등 개인정보 감지 여부", "True / False"),
    ("댓글원문", "product", "제품군", "분석 대상 제품 카테고리", "식기세척기 / 세탁기 / 냉장고 등"),
    ("댓글원문", "keyword", "검색 키워드", "이 댓글을 수집할 때 사용한 검색어", "텍스트"),
    ("댓글원문", "region", "수집 국가", "댓글 수집 대상 국가 코드", "KR / US / JP 등"),
    ("댓글원문", "title", "영상 제목", "댓글이 달린 YouTube 영상 제목", "텍스트"),
    ("댓글원문", "channel_title", "채널명", "영상이 업로드된 YouTube 채널 이름", "텍스트"),
    ("댓글원문", "video_url", "영상 URL", "원본 YouTube 영상 링크", "URL"),
    ("댓글원문", "product_related", "제품 관련 여부", "댓글이 제품 관련 내용인지 판단 결과", "True / False"),
    ("댓글원문", "product_target", "제품 대상", "언급된 구체적 제품 모델/유형", "dishwasher / washer / dryer / refrigerator 등"),
    ("댓글원문", "sentiment_label", "감성 분류", "댓글의 최종 감성 분류 (NLP 우선, 규칙 기반 fallback)", "positive / negative / neutral / mixed"),
    ("댓글원문", "sentiment_score", "감성 점수", "감성 강도 수치. 양수=긍정, 음수=부정", "-1.0 ~ 1.0"),
    ("댓글원문", "signal_strength", "이슈 신호 강도", "댓글의 비즈니스 이슈 신호 강도. nlp_sentiment_intensity 0.7 이상이면 강제 strong", "strong / medium / weak"),
    ("댓글원문", "classification_type", "분류 유형", "댓글의 내용 유형 분류", "complaint / preference / comparison / informational"),
    ("댓글원문", "confidence_level", "분류 신뢰도", "감성·유형 분류 결과에 대한 신뢰 수준", "high / medium / low"),
    ("댓글원문", "brand_label", "브랜드 언급", "댓글에서 감지된 브랜드명", "LG / Samsung / 기타 등"),
    ("댓글원문", "classification_reason", "분류 근거", "감성·유형 분류의 판단 근거 설명 (한국어)", "텍스트"),
    ("댓글원문", "insight_type", "인사이트 유형", "비즈니스 관점의 댓글 유형 분류", "품질_불만 / 만족_포인트 / 경쟁사_비교 / 사용_팁 / 문의_피드백 / 가격_민감 / 서비스_불만 / 일반_의견"),
    ("댓글원문", "parent_comment_text", "상위 댓글 원문", "대댓글인 경우 원본 상위 댓글 텍스트 (맥락 파악용)", "텍스트 또는 빈값"),
    ("댓글원문", "nlp_label", "LLM 감성 분류", "LLM(Claude/GPT)이 직접 판단한 감성 레이블", "positive / negative / neutral / trash / undecidable"),
    ("댓글원문", "nlp_confidence", "LLM 분류 신뢰도", "LLM이 반환한 분류 신뢰도 점수", "0.0 ~ 1.0"),
    ("댓글원문", "nlp_sentiment_reason", "LLM 판단 근거", "LLM이 해당 감성으로 분류한 이유 (한국어 설명)", "텍스트"),
    ("댓글원문", "nlp_is_inquiry", "문의 여부", "LLM이 판단한 질문/문의 댓글 여부", "True / False"),
    ("댓글원문", "nlp_is_rhetorical", "반어적 표현 여부", "풍자·반어법으로 작성된 댓글 여부 (오분류 방지용)", "True / False"),
    ("댓글원문", "nlp_summary", "LLM 요약", "LLM이 생성한 댓글 핵심 내용 요약 (1~2문장)", "텍스트"),
    ("댓글원문", "nlp_user_wants", "사용자 니즈", "요약에서 추출한 사용자의 구체적 wants/needs. 부정=원하는 것, 긍정=만족 포인트", "텍스트 또는 빈값"),
    ("댓글원문", "nlp_keywords", "LLM 키워드", "LLM이 추출한 핵심 단어 목록", "쉼표 구분 단어 목록"),
    ("댓글원문", "nlp_core_points", "핵심 포인트", "LLM이 추출한 댓글의 핵심 주제어 목록", "쉼표 구분 단어 목록"),
    ("댓글원문", "nlp_context_tags", "맥락 태그", "LLM이 분류한 댓글 맥락 카테고리 (복수 선택 가능)", "성능·품질 / 사용성·편의 / A/S·수리 / 가격·비용 / 불만 신호 / 강점 신호 / 문의/해결요청 / 구매 전환 / 추천 의도 등"),
    ("댓글원문", "nlp_sentiment_intensity", "LLM 감성 강도", "LLM이 판단한 감성 표현의 절대 강도. 0.7 이상이면 signal_strength strong으로 보정됨", "0.0 ~ 1.0"),
    ("댓글원문", "journey_stage", "구매 여정 단계", "댓글 작성자의 제품 구매 여정 단계 분류", "awareness / exploration / decision / purchase / delivery / onboarding / usage / maintenance / replacement"),
    ("댓글원문", "judgment_axes", "판단 축", "해당 댓글이 속한 전략적 판단 축 (최대 3개)", "price_value_fit / maintenance_burden / durability_lifespan / automation_convenience / service_after_sales / brand_trust / comparison_substitute_judgment / design_interior_fit / usage_pattern_fit / repurchase_recommendation / installation_space_fit / purchase_strategy_contract"),
    ("댓글원문", "topic_label", "토픽 레이블", "클러스터링으로 배정된 토픽 이름", "텍스트"),
    # ── 영상목록 ──────────────────────────────────────────────────────────────
    ("영상목록", "video_id", "영상 ID", "YouTube 영상 고유 식별자", "YouTube video ID"),
    ("영상목록", "title", "영상 제목", "YouTube 영상 제목", "텍스트"),
    ("영상목록", "channel_title", "채널명", "영상 업로드 채널명", "텍스트"),
    ("영상목록", "published_at", "업로드 일시", "영상이 게시된 날짜/시간", "ISO 8601 datetime"),
    ("영상목록", "view_count", "조회수", "영상 누적 조회수", "0 이상 정수"),
    ("영상목록", "like_count", "좋아요 수", "영상 누적 좋아요 수", "0 이상 정수"),
    ("영상목록", "comment_count", "댓글 수", "영상에 달린 전체 댓글 수", "0 이상 정수"),
    ("영상목록", "selection_score", "영상 선정 점수", "분석 대상 영상으로 선정될 때의 관련성 점수", "0.0 ~ 1.0"),
    ("영상목록", "video_url", "영상 URL", "YouTube 영상 직접 링크", "URL"),
    # ── 주간감성추이 ──────────────────────────────────────────────────────────
    ("주간감성추이", "week_label", "주차", "해당 주차 라벨 (예: 2025-W10)", "YYYY-Www 형식"),
    ("주간감성추이", "sentiment_label", "감성", "해당 주차의 감성 구분", "positive / negative / neutral"),
    ("주간감성추이", "ratio", "비율", "전체 유효 댓글 중 해당 감성 비율", "0.0 ~ 1.0"),
    ("주간감성추이", "comment_count", "댓글 수", "해당 주차·감성의 댓글 건수", "0 이상 정수"),
    # ── 주간키워드 ────────────────────────────────────────────────────────────
    ("주간키워드", "week_label", "주차", "해당 주차 라벨", "YYYY-Www 형식"),
    ("주간키워드", "keyword", "키워드", "해당 주차에 등장한 키워드", "텍스트"),
    ("주간키워드", "sentiment_label", "감성", "이 키워드가 주로 등장한 감성 맥락", "positive / negative / neutral"),
    ("주간키워드", "count", "등장 횟수", "해당 주차에 이 키워드가 등장한 댓글 수", "0 이상 정수"),
    # ── 긍정부정분포 ──────────────────────────────────────────────────────────
    ("긍정부정분포", "sentiment_label", "감성", "감성 분류 항목", "positive / negative / neutral"),
    ("긍정부정분포", "count", "댓글 수", "해당 감성으로 분류된 댓글 총 수", "0 이상 정수"),
    ("긍정부정분포", "ratio", "비율", "전체 유효 댓글 중 비율", "0.0 ~ 1.0"),
    # ── CEJ부정률 ─────────────────────────────────────────────────────────────
    ("CEJ부정률", "topic_label", "여정 단계", "구매 여정 단계명 (한국어)", "인지 / 탐색 / 결정 / 구매 / 배송 / 사용준비 / 사용 / 관리 / 교체"),
    ("CEJ부정률", "comments", "전체 댓글 수", "해당 여정 단계에 속한 댓글 총 수", "0 이상 정수"),
    ("CEJ부정률", "negative_comments", "부정 댓글 수", "해당 여정 단계의 부정 감성 댓글 수", "0 이상 정수"),
    ("CEJ부정률", "negative_rate", "부정 비율", "해당 여정 단계 내 부정 감성 댓글 비율. 높을수록 해당 단계에서 마찰이 많음", "0.0 ~ 1.0"),
    # ── 브랜드언급비율 ────────────────────────────────────────────────────────
    ("브랜드언급비율", "brand_label", "브랜드", "언급된 브랜드명", "LG / Samsung / 기타 등"),
    ("브랜드언급비율", "count", "언급 댓글 수", "해당 브랜드를 언급한 댓글 수", "0 이상 정수"),
    ("브랜드언급비율", "ratio", "언급 비율", "전체 유효 댓글 중 이 브랜드 언급 비율", "0.0 ~ 1.0"),
    # ── 영상당부정밀도 ────────────────────────────────────────────────────────
    ("영상당부정밀도", "title", "영상 제목", "YouTube 영상 제목", "텍스트"),
    ("영상당부정밀도", "negative_comments", "부정 댓글 수", "해당 영상의 부정 감성 댓글 수", "0 이상 정수"),
    ("영상당부정밀도", "total_comments", "전체 댓글 수", "해당 영상의 분석 대상 댓글 수", "0 이상 정수"),
    ("영상당부정밀도", "negative_density", "부정 밀도", "전체 대비 부정 댓글 비율. 높을수록 해당 영상에 불만이 집중됨", "0.0 ~ 1.0"),
    # ── 신규이슈키워드 ────────────────────────────────────────────────────────
    ("신규이슈키워드", "keyword", "키워드", "최근 처음 등장한 이슈 키워드", "텍스트"),
    ("신규이슈키워드", "latest_week", "최초 등장 주차", "이 키워드가 처음 등장한 주차", "YYYY-Www 형식"),
    ("신규이슈키워드", "count", "등장 횟수", "해당 키워드가 등장한 댓글 수", "0 이상 정수"),
    # ── 지속이슈키워드 ────────────────────────────────────────────────────────
    ("지속이슈키워드", "keyword", "키워드", "여러 주차에 걸쳐 반복 등장한 이슈 키워드", "텍스트"),
    ("지속이슈키워드", "weeks_active", "지속 주차 수", "이 키워드가 등장한 주차의 수. 많을수록 만성 이슈 가능성 높음", "1 이상 정수"),
    ("지속이슈키워드", "latest_week", "최근 등장 주차", "가장 최근에 이 키워드가 등장한 주차", "YYYY-Www 형식"),
    ("지속이슈키워드", "count", "총 등장 횟수", "전체 기간 내 이 키워드가 등장한 댓글 수", "0 이상 정수"),
    # ── 제거댓글요약 ──────────────────────────────────────────────────────────
    ("제거댓글요약", "comment_validity", "유효성", "댓글 유효성 분류 항목", "valid / excluded"),
    ("제거댓글요약", "exclusion_reason", "제외 사유", "제외된 이유 카테고리", "spam / too_short / off_topic / duplicate 등"),
    ("제거댓글요약", "count", "댓글 수", "해당 분류에 속한 댓글 수", "0 이상 정수"),
    # ── 대표댓글 (댓글원문과 동일 컬럼 설명 공유, 추가 컬럼만 기재) ─────────
    ("대표댓글", "representative_score", "대표성 점수", "이 댓글이 대표 댓글로 선정된 종합 점수", "0.0 ~ 1.0"),
    ("대표댓글", "representative_reason", "대표 선정 이유", "대표 댓글로 선정된 구체적 근거", "텍스트"),
]


class ExcelWriter:
    """Write a formatted Excel workbook with business users."""

    def __init__(self, datetime_format: str = "yyyy-mm-dd hh:mm:ss", freeze_panes: str = "A2") -> None:
        self.datetime_format = datetime_format
        self.freeze_panes = freeze_panes

    def write_workbook(self, *, workbook_path: Path, manifest: dict, videos_df: pd.DataFrame, comments_df: pd.DataFrame, analytics_outputs: dict[str, pd.DataFrame], errors_df: pd.DataFrame) -> None:
        workbook = Workbook()
        workbook.remove(workbook.active)
        workbook.properties.creator = "Codex"
        workbook.properties.title = "YouTube Comment Analysis"
        workbook.properties.subject = manifest.get("keyword", "YouTube comments")
        workbook.properties.description = "Official YouTube Data API v3 기반 분석 결과"

        self._write_readme_sheet(workbook, manifest)
        self._write_column_descriptions_sheet(workbook)
        self._write_dataframe_sheet(workbook, "핵심지표", analytics_outputs.get("runs_summary", pd.DataFrame()))
        self._write_dataframe_sheet(workbook, "영상목록", videos_df)
        self._write_dataframe_sheet(workbook, "댓글원문", self._filter_comment_columns(comments_df))
        self._write_dataframe_sheet(workbook, "영상반응순위", analytics_outputs.get("video_ranking", pd.DataFrame()))
        self._write_dataframe_sheet(workbook, "주간감성추이", analytics_outputs.get("weekly_sentiment_trend", pd.DataFrame()), sentiment=True)
        self._write_dataframe_sheet(workbook, "주간키워드", analytics_outputs.get("weekly_keyword_trend", pd.DataFrame()))
        self._write_dataframe_sheet(workbook, "긍정부정분포", analytics_outputs.get("sentiment_summary", pd.DataFrame()), sentiment=True)
        self._write_dataframe_sheet(workbook, "CEJ부정률", analytics_outputs.get("cej_negative_rate", pd.DataFrame()))
        self._write_dataframe_sheet(workbook, "브랜드언급비율", analytics_outputs.get("brand_ratio", pd.DataFrame()))
        self._write_dataframe_sheet(workbook, "영상당부정밀도", analytics_outputs.get("negative_density", pd.DataFrame()))
        self._write_dataframe_sheet(workbook, "신규이슈키워드", analytics_outputs.get("new_issue_keywords", pd.DataFrame()))
        self._write_dataframe_sheet(workbook, "지속이슈키워드", analytics_outputs.get("persistent_issue_keywords", pd.DataFrame()))
        self._write_dataframe_sheet(workbook, "제거댓글요약", analytics_outputs.get("quality_summary", pd.DataFrame()))
        self._write_dataframe_sheet(workbook, "대표댓글", self._filter_comment_columns(analytics_outputs.get("representative_comments", pd.DataFrame())), sentiment=True)
        self._write_dataframe_sheet(workbook, "오류로그", errors_df)
        self._add_weekly_bar_chart(workbook)
        self._add_keyword_donut_chart(workbook, analytics_outputs)
        workbook.save(workbook_path)

    def _write_readme_sheet(self, workbook: Workbook, manifest: dict) -> None:
        sheet = workbook.create_sheet("README")
        rows = [
            ["항목", "설명"],
            ["프로젝트", "공식 YouTube Data API v3 기반 가전 VoC 분석"],
            ["제품", manifest.get("config_json", {}).get("product")],
            ["검색 키워드", manifest.get("keyword")],
            ["국가", manifest.get("config_json", {}).get("region")],
            ["실행 ID", manifest.get("run_id")],
            ["수집 시작", manifest.get("started_at")],
            ["수집 종료", manifest.get("finished_at")],
            ["설정", str(manifest.get("config_json"))],
            ["쿼터 추정", str(manifest.get("quota_estimate"))],
            ["실제 API 호출", str(manifest.get("api_calls_json"))],
            [],
            ["시트 안내", "설명"],
            ["README", "분석 실행 메타 정보 (이 시트)"],
            ["컬럼설명", "각 시트의 컬럼 의미, 설명, 가능한 값 안내"],
            ["핵심지표", "전체 수집/분석 건수 요약 KPI"],
            ["영상목록", "수집된 YouTube 영상 목록 및 기본 정보"],
            ["댓글원문", "수집된 전체 댓글 및 LLM 분석 결과 (내부 기술 컬럼 제외)"],
            ["영상반응순위", "영상별 engagement 기준 순위"],
            ["주간감성추이", "주차별 긍정/부정/중립 댓글 비율 추이"],
            ["주간키워드", "주차별 주요 키워드 등장 빈도"],
            ["긍정부정분포", "전체 기간 감성 분포 요약"],
            ["CEJ부정률", "구매 여정(CEJ) 단계별 부정 댓글 비율"],
            ["브랜드언급비율", "브랜드별 언급 빈도 및 비율"],
            ["영상당부정밀도", "영상별 부정 댓글 밀도 (집중도 파악용)"],
            ["신규이슈키워드", "최근 처음 등장한 이슈 키워드"],
            ["지속이슈키워드", "여러 주에 걸쳐 반복 등장하는 만성 이슈 키워드"],
            ["제거댓글요약", "분석에서 제외된 댓글의 사유별 요약"],
            ["대표댓글", "감성별·토픽별 대표 댓글 및 LLM 분석 결과"],
        ]
        for row in rows:
            sheet.append(row)
        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 120
        sheet.freeze_panes = "A2"
        sheet["A1"].font = Font(bold=True)
        sheet["B1"].font = Font(bold=True)
        sheet["A13"].font = Font(bold=True)
        sheet["B13"].font = Font(bold=True)

    def _write_column_descriptions_sheet(self, workbook: Workbook) -> None:
        sheet = workbook.create_sheet("컬럼설명")
        headers = ["시트명", "컬럼명", "한국어 이름", "설명", "가능한 값"]
        sheet.append(headers)

        header_fill = PatternFill("solid", fgColor="0E5A47")
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical="center")

        current_sheet = ""
        row_num = 2
        for sheet_name, col_name, korean_name, description, possible_values in _COLUMN_DESCRIPTIONS:
            sheet.append([sheet_name, col_name, korean_name, description, possible_values])

            # 시트명이 바뀔 때 구분선 역할로 배경색 변경
            if sheet_name != current_sheet:
                fill = PatternFill("solid", fgColor="EAF4F0")
                current_sheet = sheet_name
            else:
                fill = PatternFill("solid", fgColor="FFFFFF")

            for col_idx in range(1, 6):
                cell = sheet.cell(row=row_num, column=col_idx)
                cell.fill = fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")

            row_num += 1

        sheet.column_dimensions["A"].width = 18
        sheet.column_dimensions["B"].width = 28
        sheet.column_dimensions["C"].width = 20
        sheet.column_dimensions["D"].width = 55
        sheet.column_dimensions["E"].width = 45
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:E1"

    @staticmethod
    def _filter_comment_columns(df: pd.DataFrame) -> pd.DataFrame:
        """Return only user-facing columns; preserve column order from allowlist."""
        available = [c for c in _COMMENT_EXPORT_COLUMNS if c in df.columns]
        return df[available] if available else df

    def _write_dataframe_sheet(self, workbook: Workbook, title: str, df: pd.DataFrame, sentiment: bool = False) -> None:
        sheet = workbook.create_sheet(title)
        if df.empty:
            sheet.append(["message"])
            sheet.append(["No data available"])
            return
        normalized_df = df.apply(lambda column: column.map(self._normalize_cell_value))
        sheet.append(list(normalized_df.columns))
        for row in normalized_df.fillna("").itertuples(index=False, name=None):
            sheet.append(list(row))
        sheet.freeze_panes = self.freeze_panes
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0E5A47")
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max_length + 2, 60)
        if sentiment and "sentiment_label" in normalized_df.columns:
            sentiment_col = list(normalized_df.columns).index("sentiment_label") + 1
            col_letter = get_column_letter(sentiment_col)
            sheet.conditional_formatting.add(f"{col_letter}2:{col_letter}{sheet.max_row}", CellIsRule(operator="equal", formula=['"positive"'], fill=PatternFill("solid", fgColor="D7F5E8")))
            sheet.conditional_formatting.add(f"{col_letter}2:{col_letter}{sheet.max_row}", CellIsRule(operator="equal", formula=['"negative"'], fill=PatternFill("solid", fgColor="FCE0DD")))
            sheet.conditional_formatting.add(f"{col_letter}2:{col_letter}{sheet.max_row}", CellIsRule(operator="equal", formula=['"neutral"'], fill=PatternFill("solid", fgColor="F6F0D8")))

    @staticmethod
    def _normalize_cell_value(value):
        if hasattr(value, "tolist") and not isinstance(value, (str, bytes)):
            value = value.tolist()
        if isinstance(value, (list, tuple, set)):
            return ", ".join(str(item) for item in value)
        if isinstance(value, dict):
            return json.dumps(value, ensure_ascii=False)
        return value

    def _add_weekly_bar_chart(self, workbook: Workbook) -> None:
        sheet_name = "주간감성추이"
        if sheet_name not in workbook.sheetnames:
            return
        source = workbook[sheet_name]
        if source.max_row < 2:
            return
        chart_sheet = workbook.create_sheet("주간감성차트")
        chart_sheet.append(["주차", "긍정 응답률", "부정 응답률"])
        headers = [cell.value for cell in source[1]]
        week_idx = headers.index("week_label") + 1 if "week_label" in headers else None
        sentiment_idx = headers.index("sentiment_label") + 1 if "sentiment_label" in headers else None
        ratio_idx = headers.index("ratio") + 1 if "ratio" in headers else None
        if not all([week_idx, sentiment_idx, ratio_idx]):
            return
        rows: dict[str, dict[str, float]] = {}
        for row_idx in range(2, source.max_row + 1):
            week_label = source.cell(row=row_idx, column=week_idx).value
            sentiment = source.cell(row=row_idx, column=sentiment_idx).value
            ratio = source.cell(row=row_idx, column=ratio_idx).value
            if not week_label or sentiment not in {"positive", "negative"}:
                continue
            rows.setdefault(str(week_label), {"positive": 0.0, "negative": 0.0})
            rows[str(week_label)][str(sentiment)] = float(ratio)
        for week_label, values in rows.items():
            chart_sheet.append([week_label, values.get("positive", 0.0), values.get("negative", 0.0)])
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "주차별 긍정/부정 응답률"
        chart.y_axis.title = "응답률"
        chart.x_axis.title = "주차"
        data = Reference(chart_sheet, min_col=2, min_row=1, max_col=3, max_row=chart_sheet.max_row)
        cats = Reference(chart_sheet, min_col=1, min_row=2, max_row=chart_sheet.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 16
        chart_sheet.add_chart(chart, "E2")

    def _add_keyword_donut_chart(self, workbook: Workbook, analytics_outputs: dict[str, pd.DataFrame]) -> None:
        keyword_df = analytics_outputs.get("weekly_keyword_trend", pd.DataFrame())
        if keyword_df.empty:
            return
        donut_sheet = workbook.create_sheet("키워드도넛차트")
        for offset, sentiment in enumerate(["negative", "positive"]):
            subset = keyword_df[keyword_df["sentiment_label"] == sentiment].groupby("keyword", as_index=False)["count"].sum().sort_values("count", ascending=False).head(6)
            if subset.empty:
                continue
            start_row = 1 if offset == 0 else 20
            donut_sheet.cell(row=start_row, column=1, value="키워드")
            donut_sheet.cell(row=start_row, column=2, value="비중")
            for idx, row in enumerate(subset.itertuples(index=False), start=1):
                donut_sheet.cell(row=start_row + idx, column=1, value=row.keyword)
                donut_sheet.cell(row=start_row + idx, column=2, value=row.count)
            chart = DoughnutChart()
            chart.title = "부정 키워드 비중" if sentiment == "negative" else "긍정 키워드 비중"
            labels = Reference(donut_sheet, min_col=1, min_row=start_row + 1, max_row=start_row + len(subset))
            data = Reference(donut_sheet, min_col=2, min_row=start_row, max_row=start_row + len(subset))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            chart.height = 7
            chart.width = 8
            donut_sheet.add_chart(chart, "D2" if sentiment == "negative" else "D20")
