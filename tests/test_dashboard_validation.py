from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from src.dashboard_app import (
    COL_BRAND,
    COL_CEJ,
    COL_COUNTRY,
    COL_SENTIMENT,
    build_raw_download_package,
    compute_filtered_bundle,
    localize_region,
)


EXPECTED_SHEETS = {
    "댓글원문",
    "영상목록",
    "주간감성",
    "CEJ부정률",
    "브랜드언급",
    "영상밀도분석",
}


def test_build_raw_download_package_contains_expected_sheets():
    frame = pd.DataFrame([{"value": 1}])
    payload = build_raw_download_package(frame, frame, frame, frame, frame, frame)
    workbook = load_workbook(BytesIO(payload))

    assert EXPECTED_SHEETS.issubset(set(workbook.sheetnames))


def test_compute_filtered_bundle_supports_inquiry_scope_and_empty_videos():
    comments = pd.DataFrame(
        [
            {
                "comment_id": "c1",
                "product": "dryer",
                "region": "KR",
                COL_COUNTRY: localize_region("KR"),
                COL_BRAND: "LG",
                COL_CEJ: "기타",
                COL_SENTIMENT: "중립",
                "text_display": "설치 어떻게 하나요",
                "cleaned_text": "설치 어떻게 하나요",
                "comment_validity": "valid",
            },
            {
                "comment_id": "c2",
                "product": "dryer",
                "region": "KR",
                COL_COUNTRY: localize_region("KR"),
                COL_BRAND: "LG",
                COL_CEJ: "기타",
                COL_SENTIMENT: "부정",
                "text_display": "소음이 너무 심해요",
                "cleaned_text": "소음이 너무 심해요",
                "comment_validity": "valid",
            },
        ]
    )
    opinion_units = pd.DataFrame(
        [
            {
                "source_content_id": "c1",
                "inquiry_flag": True,
                "product": "dryer",
                "region": "KR",
                "brand_mentioned": "",
                "cej_scene_code": "S10 Other",
                "sentiment": "neutral",
            },
            {
                "source_content_id": "c2",
                "inquiry_flag": False,
                "product": "dryer",
                "region": "KR",
                "brand_mentioned": "",
                "cej_scene_code": "S10 Other",
                "sentiment": "negative",
            },
        ]
    )

    bundle = compute_filtered_bundle(
        comments,
        pd.DataFrame(),
        pd.DataFrame(),
        {
            "products": ["dryer"],
            "regions": [localize_region("KR")],
            "sentiments": [],
            "cej": [],
            "brands": [],
            "keyword_query": "",
            "analysis_scope": "문의 포함 댓글만",
        },
        pd.DataFrame(),
        opinion_units,
    )

    assert bundle["inquiry_count"] == 1
    assert list(bundle["comments"]["comment_id"]) == ["c1"]
