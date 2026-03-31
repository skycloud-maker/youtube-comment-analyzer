from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.config import load_config
from src.dashboard_app import (
    add_localized_columns,
    build_raw_download_package,
    build_weekly_sentiment_window,
    compute_filtered_bundle,
    load_dashboard_data,
    localize_region,
    _build_dashboard_options,
    _latest_signature,
    _read_cache_bundle,
)

BASE_DIR = Path(__file__).resolve().parents[1]
REPORT_PATH = BASE_DIR / "docs" / "harness" / "generated" / "dashboard-validation-report.md"
EXPECTED_SHEETS = {
    "댓글원문",
    "영상목록",
    "주간감성",
    "CEJ부정률",
    "브랜드언급",
    "영상당부정밀도",
}


def _region_codes_from_labels(values: list[str]) -> list[str]:
    mapping = {
        localize_region("KR"): "KR",
        localize_region("US"): "US",
        "KR": "KR",
        "US": "US",
    }
    return [mapping.get(value, value) for value in values]


def _filter_aux_frames(data: dict[str, pd.DataFrame], selected_filters: dict[str, object]) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    products = selected_filters.get("products") or []
    region_codes = _region_codes_from_labels(selected_filters.get("regions") or [])

    cej_df = data.get("cej_negative_rate", pd.DataFrame()).copy()
    if not cej_df.empty:
        if products and "product" in cej_df.columns:
            cej_df = cej_df[cej_df["product"].isin(products)]
        if region_codes and "region" in cej_df.columns:
            cej_df = cej_df[cej_df["region"].isin(region_codes)]

    brand_df = data.get("brand_ratio", pd.DataFrame()).copy()
    if not brand_df.empty:
        if products and "product" in brand_df.columns:
            brand_df = brand_df[brand_df["product"].isin(products)]
        if region_codes and "region" in brand_df.columns:
            brand_df = brand_df[brand_df["region"].isin(region_codes)]

    density_df = data.get("negative_density", pd.DataFrame()).copy()
    if not density_df.empty:
        if products and "product" in density_df.columns:
            density_df = density_df[density_df["product"].isin(products)]
        if region_codes and "region" in density_df.columns:
            density_df = density_df[density_df["region"].isin(region_codes)]

    return cej_df, brand_df, density_df


def _build_filter_scenarios(options: dict[str, list[str]]) -> list[dict[str, object]]:
    base_products = options["available_products"][: min(2, len(options["available_products"]))] or []
    base_regions = options["regions"][:]
    scenarios: list[dict[str, object]] = [
        {
            "products": base_products,
            "regions": base_regions,
            "sentiments": [],
            "cej": [],
            "brands": [],
            "keyword_query": "",
            "analysis_scope": "전체",
        },
        {
            "products": base_products[:1],
            "regions": base_regions[:1],
            "sentiments": ["부정"],
            "cej": [],
            "brands": [],
            "keyword_query": "",
            "analysis_scope": "전체",
        },
        {
            "products": base_products[:1],
            "regions": base_regions[-1:],
            "sentiments": ["긍정"],
            "cej": [],
            "brands": [],
            "keyword_query": "",
            "analysis_scope": "전체",
        },
        {
            "products": base_products[:1],
            "regions": base_regions[:1],
            "sentiments": [],
            "cej": [],
            "brands": [],
            "keyword_query": "",
            "analysis_scope": "문의 포함 댓글만",
        },
        {
            "products": base_products[:1],
            "regions": base_regions[:1],
            "sentiments": [],
            "cej": [],
            "brands": [],
            "keyword_query": "",
            "analysis_scope": "문의 제외",
        },
        {
            "products": base_products,
            "regions": base_regions,
            "sentiments": [],
            "cej": [],
            "brands": [],
            "keyword_query": "warranty",
            "analysis_scope": "전체",
        },
    ]
    return scenarios


def main() -> None:
    results: list[tuple[str, str]] = []
    failures: list[str] = []

    try:
        config = load_config()
        processed_dir = config.yaml.storage.processed_dir
        raw_dir = config.yaml.storage.raw_dir
        results.append(("config", f"processed={processed_dir}, raw={raw_dir}"))
    except Exception as exc:
        failures.append(f"config load failed: {exc}")
        processed_dir = BASE_DIR / "data" / "processed"
        raw_dir = BASE_DIR / "data" / "raw"

    data = load_dashboard_data()
    comments_df = add_localized_columns(data.get("comments", pd.DataFrame()))
    videos_df = data.get("videos", pd.DataFrame()).copy()
    options = _build_dashboard_options(comments_df, videos_df)

    if comments_df.empty:
        failures.append("dashboard comments dataset is empty")
    else:
        results.append(("data", f"comments={len(comments_df):,}, videos={len(videos_df):,}"))

    latest_signature = _latest_signature()
    cache_path = BASE_DIR / "data" / "dashboard_cache" / "dashboard_bundle.pkl"
    cached = _read_cache_bundle(cache_path, latest_signature)
    if cached is None:
        fallback_cache = _read_cache_bundle(cache_path, None)
        if fallback_cache is not None:
            results.append(("cache", "stale or signature-mismatched cache detected; loader fallback remains available"))
        else:
            results.append(("cache", "no reusable cache detected; dataset loader fallback remains available"))
    else:
        results.append(("cache", "latest signature cache read succeeded"))

    raw_json_count = len(list(Path(raw_dir).rglob("*.json")))
    if raw_json_count == 0:
        failures.append("no raw YouTube audit json files found")
    else:
        results.append(("raw_audit", f"json_files={raw_json_count:,}"))

    scenarios = _build_filter_scenarios(options)
    for index, filters in enumerate(scenarios, start=1):
        try:
            bundle = compute_filtered_bundle(
                comments_df,
                videos_df,
                data.get("quality_summary", pd.DataFrame()),
                filters,
                data.get("representative_bundles", pd.DataFrame()),
                data.get("opinion_units", pd.DataFrame()),
            )
            weekly_window, _ = build_weekly_sentiment_window(bundle["comments"], weeks_per_view=8, page=1)
            cej_df, brand_df, density_df = _filter_aux_frames(data, filters)
            export_bytes = build_raw_download_package(bundle["comments"], bundle["videos"], weekly_window, cej_df, brand_df, density_df)
            workbook = load_workbook(BytesIO(export_bytes))
            if not EXPECTED_SHEETS.issubset(set(workbook.sheetnames)):
                raise AssertionError(f"missing sheets: {EXPECTED_SHEETS - set(workbook.sheetnames)}")
        except Exception as exc:
            failures.append(f"filter scenario {index} failed: {filters} -> {exc}")

    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "# Dashboard Validation Report",
        "",
        "## Summary",
        f"- scenarios checked: {len(scenarios)}",
        f"- failures: {len(failures)}",
        "",
        "## Passed Checks",
    ]
    lines.extend([f"- `{name}`: {detail}" for name, detail in results])
    lines.append("")
    lines.append("## Failures")
    if failures:
        lines.extend([f"- {failure}" for failure in failures])
    else:
        lines.append("- none")
    REPORT_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")

    payload = {"failures": failures, "report": str(REPORT_PATH), "scenario_count": len(scenarios)}
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    if failures:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
